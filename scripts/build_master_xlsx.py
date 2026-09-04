#!/usr/bin/env python3
"""Master provider call sheet — every TDHCA organization, full script capture columns.

One row per organization (one call covers all their programs). Navy columns are call
context; crimson/yellow columns are what the VA fills in, in the order
docs/05-va-operations.md asks for them. The WEEK 1 flag marks the 16 CEAP providers
serving the ten launch counties — start there.

  python3 scripts/build_master_xlsx.py [src.csv] [out.xlsx]
"""
import csv, sys, json, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else 'data/call-sheets/master-provider-list.csv'
OUT = sys.argv[2] if len(sys.argv) > 2 else 'data/call-sheets/CornerHelp-Master-Provider-List.xlsx'
LISTINGS = 'data/listings/tdhca-subrecipients.jsonl'
LAUNCH = {'Harris','Dallas','Tarrant','Bexar','Travis','Collin','Denton','Hidalgo','El Paso','Fort Bend'}

NAVY='1C2B5E'; CRIMSON='C41F4E'; LAV='F1F4FC'; LINE='DFE4F2'; YELLOW='FFF9D9'; GREY='F5F6FA'

rows = list(csv.DictReader(open(SRC, encoding='utf-8')))
week1 = set()
if pathlib.Path(LISTINGS).exists():
    for line in open(LISTINGS, encoding='utf-8'):
        if not line.strip(): continue
        r = json.loads(line)
        if 'CEAP' in r['program_name'] and LAUNCH & set(r['service_counties']):
            week1.add(r['org_name'])

CONTEXT = [('week1','WEEK 1',9),('rank','#',5),('org_name','Organization',40),
           ('city','City',14),('address','Address',30),('phone','Phone',15),
           ('programs','Programs',17),('needs','Help offered',36),
           ('counties_count','Counties',9),('zips_count','ZIPs',8),
           ('population_reach','People reached',13),
           ('counties','Counties served',55),('zips','ZIP codes served',55),
           ('source','Source',34)]
CAPTURE = [('status','Status',18),
           # Ask in the rep's vocabulary, not ours. Nobody at a food pantry can recite a
           # ZIP list; they can tell you whether they cover the whole county and where
           # they turn people away. We convert that to ZIPs — that is our job, not theirs.
           ('service_area_stated','SERVICE AREA — whole county? which areas?',32),
           ('funding_lasts_until','Funding lasts until',21),
           ('reopens_on','Reopens on',13),('how_to_apply','How to apply',14),
           ('hours','Hours',24),('daily_cap','Daily cap',10),
           ('documents_required','Documents required',38),
           ('most_common_turnaway','Most common reason turned away',44),
           ('anything_changing','Anything changing',28),('spoke_with','Spoke with',17),
           ('verify_outcome','Call outcome',16),('note','Note',36),
           ('called_at','Called (date)',13),('va','VA',11)]
ALL = CONTEXT + CAPTURE
capkeys = {k for k,_,_ in CAPTURE}

wb = Workbook(); ws = wb.active; ws.title = 'Master List'
thin = Side(style='thin', color=LINE); bd = Border(left=thin,right=thin,top=thin,bottom=thin)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ALL))
b = ws.cell(1,1,f'CornerHelp — master provider call sheet · {len(rows)} organizations · all 254 Texas counties')
b.font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
b.fill = PatternFill('solid', fgColor=NAVY); b.alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 26

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(CONTEXT))
n = ws.cell(2,1,'DO NOT EDIT — call context')
n.font = Font(name='Arial', size=9, bold=True, color='6B78A0')
n.fill = PatternFill('solid', fgColor=GREY); n.alignment = Alignment(horizontal='center')
ws.merge_cells(start_row=2, start_column=len(CONTEXT)+1, end_row=2, end_column=len(ALL))
n = ws.cell(2, len(CONTEXT)+1,
            'FILL THESE IN — in the order the script asks for them. Blank beats guessed.')
n.font = Font(name='Arial', size=9, bold=True, color='8A6D1F')
n.fill = PatternFill('solid', fgColor=YELLOW); n.alignment = Alignment(horizontal='center')

HR = 3
for i,(k,label,w) in enumerate(ALL, 1):
    c = ws.cell(HR, i, label)
    c.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=CRIMSON if k in capkeys else NAVY)
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    c.border = bd
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[HR].height = 34

for r,row in enumerate(rows, HR+1):
    for i,(k,label,w) in enumerate(ALL, 1):
        if k == 'week1':
            v = 'WEEK 1' if row['org_name'] in week1 else ''
        else:
            v = row.get(k,'')
            if k in ('rank','counties_count','zips_count','population_reach'):
                v = int(v) if str(v).strip() else ''
        c = ws.cell(r, i, v)
        c.font = Font(name='Arial', size=10,
                      bold=(k=='week1' and v!=''), color=CRIMSON if k=='week1' and v else '000000')
        c.border = bd
        c.alignment = Alignment(vertical='top', wrap_text=k in ('org_name','counties','zips','needs','address','source'),
                                horizontal='center' if k=='week1' else None)
        if k in capkeys: c.fill = PatternFill('solid', fgColor=YELLOW)
        elif r % 2 == 0: c.fill = PatternFill('solid', fgColor=LAV)
        if k == 'population_reach': c.number_format = '#,##0'
    ws.row_dimensions[r].height = 30

last = HR + len(rows)
def dv(formula, key, prompt):
    d = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
    d.prompt = prompt; d.showInputMessage = True
    ws.add_data_validation(d)
    L = get_column_letter([k for k,_,_ in ALL].index(key)+1)
    d.add(f'{L}{HR+1}:{L}{last}')
dv('"accepting,waitlist,funds_exhausted,seasonal_closed,appointment_only,unknown"','status',
   'Never a yes/no question. Ask: "How far into the month is the funding usually lasting right now?"')
dv('"walk_in,phone,online,appointment"','how_to_apply','Read the prefilled fact back as a statement.')
dv('"reached,voicemail,no_answer,busy,wrong_number,callback_booked,refused,gatekeeper,disconnected"',
   'verify_outcome','Log every attempt, including the ones that fail.')

ws.freeze_panes = f'D{HR+1}'
ws.auto_filter.ref = f'A{HR}:{get_column_letter(len(ALL))}{last}'

# ---------- how to use ----------
gs = wb.create_sheet('How to use')
gs.column_dimensions['A'].width = 30; gs.column_dimensions['B'].width = 94
gs.merge_cells('A1:B1')
t = gs.cell(1,1,'How to use this sheet'); t.font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
t.fill = PatternFill('solid', fgColor=NAVY); gs.row_dimensions[1].height = 24

def line(r, a, b='', fill=None, size=10):
    ca = gs.cell(r,1,a); cb = gs.cell(r,2,b)
    ca.font = Font(name='Arial', size=size, bold=True, color='1C2B5E')
    cb.font = Font(name='Arial', size=size, color='1C2B5E')
    cb.alignment = Alignment(wrap_text=True, vertical='top')
    if fill:
        ca.fill = PatternFill('solid', fgColor=fill); cb.fill = PatternFill('solid', fgColor=fill)
    return r+1

r = 3
r = line(r,'Start with WEEK 1',f'{len(week1)} organizations flagged in column A. They are the CEAP '
        'providers serving the ten launch counties, and between them they cover 68 counties — '
        '74% of the state. Filter column A to WEEK 1 and work that set first.')
r = line(r,'One row = one call','An organization runs several programs (CEAP, CSBG, WAP). One call '
        'covers all of them. The Programs column says which apply.')
r = line(r,'Counties vs ZIPs','Counties are what the source declared. ZIP codes are the crosswalk '
        'expansion and are what the site actually searches on — HUD residential-share data, '
        'marginal slivers under 5% dropped. If an agency tells you their service area is '
        'different from what is listed, say so in Note; that correction is worth more than any '
        'other field on the row.')
r = line(r,'Which cells to edit','Only the YELLOW columns, from Status rightward. Navy columns are '
        'call context — leave them alone; they are how the row matches back to the database.')
r = line(r,'Service area — ask it their way','Do NOT ask "which ZIP codes do you serve" — '
        'no intake worker can answer that. Ask: "Do you cover all of [county], or just part '
        'of it?" and then, if part: "Where do you have to turn people away for being outside '
        'your area?" Write down exactly what they say — "Spring Branch", "inside 610", '
        '"anywhere in the county". Converting that to ZIP codes is our job.')
r = line(r,'Some rows need no answer','Agencies running CEAP, CSBG or WAP have their counties '
        'set by their TDHCA contract — the state publishes it and their answer cannot improve '
        'on it. Those rows are already publishable; skip the question. It matters for the '
        'pantries, churches and ministries, where a county is a guess.')
r = line(r,'Blank beats guessed','If it was not said, leave it empty and put the reason in Note '
        '("refused", "did not know", "ran out of time"). A guessed field poisons the freshness '
        'claim, which is the whole product. VAs are measured on accuracy, never on how many cells '
        'are full.')
r = line(r,'Verify the row first','Every row came from a model reading a PDF. Before trusting the '
        'name or counties, confirm the phone actually reaches that organization. If it does not, '
        'set Call outcome to wrong_number and stop.')
r = line(r,'Three attempts','Different days, different times. Log every attempt including voicemail '
        'and no-answer. Then leave it — an honest "we could not reach them" beats a stale green.')
r += 1
r = line(r,'THE CALL',' ', fill=LAV)
for step, txt in [
 ('1 · Open','"Hi, this is [name] — I am calling from CornerHelp, we are the site that lists which '
             'programs are taking applications. We send people to you. I have about two minutes of '
             'questions so we are not sending you folks you cannot help. Is now alright, or should '
             'I call back?"'),
 ('2 · Status','"How far into the month is the funding usually lasting right now?" — never a yes/no '
               'question. Captures Status, Funding lasts until, Reopens on.'),
 ('3 · Practicals','Read the prefilled facts back as statements: "I have you as walk-in, Monday to '
                   'Thursday, doors at eight, capping around 40 a day. Still right?" Captures How to '
                   'apply, Hours, Daily cap, Documents required.'),
 ('4 · The money question','"What is the most common reason somebody gets turned away here?" Record '
                           'it VERBATIM, in their words. Highest-value field on the sheet and no '
                           'competitor can scrape it.'),
 ('5 · Read back','Say the record back out loud, then: "I will put today\'s date on it so people know '
                  'it is current. If anything changes, there is a link on your listing to update it '
                  'yourself."')]:
    r = line(r, step, txt)
r += 1
r = line(r,'NEVER',' ', fill=LAV)
r = line(r,'Do not say','Partner. Work with. Affiliated with. On behalf of. We are with the state. '
        'Calling for a client.')
r = line(r,'Do not do','Claim a partnership or referral agreement · ask for client names or case data '
        '· offer anything of value for information · promise traffic or leads · negotiate placement · '
        'argue with a removal request. Any of these is a same-day retrain.')
r = line(r,'Removal request','Honor it same day, no second ask. Log it. That is the reputation the '
        'whole operation runs on.')
r += 1
r = line(r,'EXAMPLE ROW',' ', fill=LAV)
r = line(r,'','A worked example of the expected format. It lives here, not on the Master List, so it '
        'can never be mistaken for data and imported as a real listing.')
for k,v in [('Status','accepting'),('Funding lasts until','around the 8th–10th'),
            ('Reopens on','2026-10-01'),('How to apply','walk_in'),('Hours','Mon–Thu, doors at 8:00'),
            ('Daily cap','40'),
            ('Documents required','photo ID; SSN cards all household members; current electric bill; '
                                  '30 days income all adults'),
            ('Most common reason turned away','"People come in without the income for everybody in '
                                              'the house"'),
            ('SERVICE AREA — whole county? which areas?','not the whole county — west side, roughly Spring Branch out to Katy'),
            ('Anything changing','new intake system starting November'),
            ('Spoke with','Denise, intake'),('Call outcome','reached'),
            ('Called (date)','2026-09-03'),('VA','JM')]:
    ca = gs.cell(r,1,k); cb = gs.cell(r,2,v)
    ca.font = Font(name='Arial', size=9, color='6B78A0'); cb.font = Font(name='Arial', size=9)
    cb.fill = PatternFill('solid', fgColor=YELLOW)
    cb.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
r += 1
r = line(r,'PROGRESS',' ', fill=LAV)
L = get_column_letter([k for k,_,_ in ALL].index('verify_outcome')+1)
S = get_column_letter([k for k,_,_ in ALL].index('status')+1)
for label, formula, fmt in [
    ('Calls logged', f"=COUNTA('Master List'!{L}{HR+1}:{L}{last})&\" of {len(rows)}\"", None),
    ('Reached',      f"=COUNTIF('Master List'!{L}{HR+1}:{L}{last},\"reached\")", None),
    ('Currently accepting', f"=COUNTIF('Master List'!{S}{HR+1}:{S}{last},\"accepting\")", None),
    ('Reach rate',   f"=IFERROR(COUNTIF('Master List'!{L}{HR+1}:{L}{last},\"reached\")"
                     f"/COUNTA('Master List'!{L}{HR+1}:{L}{last}),\"—\")", '0%')]:
    gs.cell(r,1,label).font = Font(name='Arial', size=10, bold=True, color=NAVY)
    c = gs.cell(r,2,formula); c.font = Font(name='Arial', size=10, bold=(fmt=='0%'))
    if fmt: c.number_format = fmt
    r += 1
gs.cell(r,2,'This is the number the staffing model is guessing at (it assumes 60%). Measure it here.')\
  .font = Font(name='Arial', size=9, italic=True, color='6B78A0')

wb.active = 0
wb.save(OUT); print(f'wrote {OUT} · {len(rows)} orgs · {len(week1)} flagged WEEK 1 · '
                    f'{len(CAPTURE)} capture columns')
