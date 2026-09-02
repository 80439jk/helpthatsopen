import csv, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

import sys
SRC=sys.argv[1] if len(sys.argv)>1 else 'data/call-sheets/week-01-ceap-launch-counties.csv'
rows=list(csv.DictReader(open(SRC, encoding='utf-8')))

NAVY='1C2B5E'; LAV='F1F4FC'; LINE='DFE4F2'; CRIMSON='C41F4E'
YELLOW='FFF9D9'; GREY='F5F6FA'

CONTEXT=['rank','org_name','program','phone','city','counties','population_reach','largest_counties']
CAPTURE=['status','funding_lasts_until','reopens_on','how_to_apply','hours','daily_cap',
         'documents_required','most_common_turnaway','anything_changing','spoke_with',
         'verify_outcome','note','called_at','va']
HEADERS={'rank':'#','org_name':'Organization','program':'Program','phone':'Phone','city':'City',
 'counties':'Counties','population_reach':'People reached','largest_counties':'Largest counties',
 'status':'Status','funding_lasts_until':'Funding lasts until','reopens_on':'Reopens on',
 'how_to_apply':'How to apply','hours':'Hours','daily_cap':'Daily cap',
 'documents_required':'Documents required','most_common_turnaway':'Most common reason turned away',
 'anything_changing':'Anything changing','spoke_with':'Spoke with','verify_outcome':'Call outcome',
 'note':'Note','called_at':'Called (date)','va':'VA'}
WIDTH={'rank':5,'org_name':44,'program':34,'phone':16,'city':15,'counties':10,
 'population_reach':14,'largest_counties':30,'status':18,'funding_lasts_until':22,
 'reopens_on':14,'how_to_apply':15,'hours':26,'daily_cap':11,'documents_required':40,
 'most_common_turnaway':46,'anything_changing':30,'spoke_with':18,'verify_outcome':17,
 'note':40,'called_at':14,'va':12}

wb=Workbook(); ws=wb.active; ws.title='Call Sheet'
thin=Side(style='thin', color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)

# banner
cols=CONTEXT+CAPTURE
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
b=ws.cell(1,1,'CornerHelp — Week 1 verification calls · CEAP · ten launch counties')
b.font=Font(name='Arial',size=13,bold=True,color='FFFFFF'); b.fill=PatternFill('solid',fgColor=NAVY)
b.alignment=Alignment(vertical='center'); ws.row_dimensions[1].height=26

ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(CONTEXT))
n=ws.cell(2,1,'DO NOT EDIT — call context'); n.font=Font(name='Arial',size=9,bold=True,color='6B78A0')
n.fill=PatternFill('solid',fgColor=GREY); n.alignment=Alignment(horizontal='center')
ws.merge_cells(start_row=2,start_column=len(CONTEXT)+1,end_row=2,end_column=len(cols))
n=ws.cell(2,len(CONTEXT)+1,'FILL THESE IN — in the order the script asks for them. Blank beats guessed.')
n.font=Font(name='Arial',size=9,bold=True,color='8A6D1F'); n.fill=PatternFill('solid',fgColor=YELLOW)
n.alignment=Alignment(horizontal='center')

HR=3
for i,c in enumerate(cols,1):
    h=ws.cell(HR,i,HEADERS[c])
    h.font=Font(name='Arial',size=10,bold=True,color='FFFFFF')
    h.fill=PatternFill('solid',fgColor=NAVY if c in CONTEXT else CRIMSON)
    h.alignment=Alignment(wrap_text=True,vertical='center',horizontal='center')
    h.border=border
    ws.column_dimensions[get_column_letter(i)].width=WIDTH[c]
ws.row_dimensions[HR].height=34

for r,row in enumerate(rows, HR+1):
    for i,c in enumerate(cols,1):
        v=row.get(c,'')
        if c in ('rank','counties','population_reach'):
            v=int(v) if str(v).strip() else ''
        cell=ws.cell(r,i,v)
        cell.font=Font(name='Arial',size=10)
        cell.border=border
        cell.alignment=Alignment(vertical='top',wrap_text=c in ('org_name','program','largest_counties'))
        if c in CAPTURE: cell.fill=PatternFill('solid',fgColor=YELLOW)
        elif r%2==0:     cell.fill=PatternFill('solid',fgColor=LAV)
        if c=='population_reach': cell.number_format='#,##0'
        if c=='phone': cell.alignment=Alignment(vertical='top')
    ws.row_dimensions[r].height=30

last=HR+len(rows)
def dv(formula, colname, prompt):
    d=DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
    d.promptTitle=HEADERS[colname]; d.prompt=prompt; d.showInputMessage=True
    ws.add_data_validation(d)
    L=get_column_letter(cols.index(colname)+1)
    d.add(f'{L}{HR+1}:{L}{last}')
dv('"accepting,waitlist,funds_exhausted,seasonal_closed,appointment_only,unknown"','status',
   'Never ask a yes/no question. Ask: "How far into the month is the funding usually lasting right now?"')
dv('"walk_in,phone,online,appointment"','how_to_apply','Read the prefilled fact back as a statement.')
dv('"reached,voicemail,no_answer,busy,wrong_number,callback_booked,refused,gatekeeper,disconnected"',
   'verify_outcome','Log every attempt, including the ones that fail.')

ws.freeze_panes=f'C{HR+1}'
ws.auto_filter.ref=f'A{HR}:{get_column_letter(len(cols))}{last}'

# ---------- instructions ----------
gs=wb.create_sheet('How to use')
gs.column_dimensions['A'].width=30; gs.column_dimensions['B'].width=94
def line(r,a,b='',bold=False,size=10,color='1C2B5E',fill=None):
    ca=gs.cell(r,1,a); cb=gs.cell(r,2,b)
    ca.font=Font(name='Arial',size=size,bold=True,color=color)
    cb.font=Font(name='Arial',size=size,bold=bold,color='1C2B5E')
    cb.alignment=Alignment(wrap_text=True,vertical='top')
    if fill:
        ca.fill=PatternFill('solid',fgColor=fill); cb.fill=PatternFill('solid',fgColor=fill)
    gs.row_dimensions[r].height=None if len(str(b))<95 else 30
    return r+1

r=1
gs.merge_cells('A1:B1')
t=gs.cell(1,1,'How to use this sheet'); t.font=Font(name='Arial',size=13,bold=True,color='FFFFFF')
t.fill=PatternFill('solid',fgColor=NAVY); gs.row_dimensions[1].height=24
r=3
r=line(r,'What this is','16 CEAP providers serving the ten launch counties — 18.4M people, 59% of Texas. Ranked by the cold-start priority score: nothing is verified yet, so the score is reach + volatility. Harris County sits at the top because it should.')
r=line(r,'Which cells to edit','Only the YELLOW columns, from "Status" rightward. The navy columns to the left are call context — leave them alone; they are how the row matches back to the database.')
r=line(r,'Blank beats guessed','If it was not said, leave the cell empty and put the reason in Note ("refused", "did not know", "ran out of time"). A guessed field poisons the freshness claim, which is the whole product. VAs are measured on accuracy, never on how many cells are full.')
r=line(r,'Verify the row first','Every row came from a model reading a PDF. Before trusting the name or counties, confirm the phone actually reaches that organization. If it does not, set Call outcome to wrong_number and stop.')
r=line(r,'Three attempts','Different days, different times. Log every attempt including voicemail and no-answer. Then leave it — an honest "we could not reach them" beats a stale green status.')
r+=1
r=line(r,'THE CALL',' ',bold=True,fill=LAV)
for step,txt in [
 ('1 · Open','"Hi, this is [name] — I am calling from CornerHelp, we are the site that lists which programs are taking applications. We send people to you. I have about two minutes of questions so we are not sending you folks you cannot help. Is now alright, or should I call back?"'),
 ('2 · Status','"How far into the month is the funding usually lasting right now?" — never a yes/no question. Captures Status, Funding lasts until, Reopens on.'),
 ('3 · Practicals','Read the prefilled facts back as statements: "I have you as walk-in, Monday to Thursday, doors at eight, capping around 40 a day. Still right?" Captures How to apply, Hours, Daily cap, Documents required.'),
 ('4 · The money question','"What is the most common reason somebody gets turned away here?" Record it VERBATIM, in their words. This is the highest-value field on the sheet and no competitor can scrape it.'),
 ('5 · Read back','Say the whole record back out loud, then: "I will put today\'s date on it so people know it is current. If anything changes, there is a link on your listing to update it yourself."')]:
    r=line(r,step,txt)
r+=1
r=line(r,'NEVER',' ',bold=True,fill=LAV)
r=line(r,'Do not say','Partner. Work with. Affiliated with. On behalf of. We are with the state. Calling for a client.')
r=line(r,'Do not do','Claim a partnership or referral agreement · ask for client names or case data · offer anything of value for information · promise traffic or leads · negotiate placement · argue with a removal request. Any of these is a same-day retrain.')
r=line(r,'Removal request','Honor it same day, no second ask. Log it. That is the reputation the whole operation runs on.')
r+=1
r=line(r,'EXAMPLE ROW',' ',bold=True,fill=LAV)
r=line(r,'','A worked example of the expected format. It lives here, not on the Call Sheet, so it can never be mistaken for data and imported as a real listing.')
ex=[('Status','accepting'),('Funding lasts until','around the 8th–10th'),('Reopens on','2026-10-01'),
    ('How to apply','walk_in'),('Hours','Mon–Thu, doors at 8:00'),('Daily cap','40'),
    ('Documents required','photo ID; SSN cards all household members; current electric bill; 30 days income all adults'),
    ('Most common reason turned away','"People come in without the income for everybody in the house"'),
    ('Anything changing','new intake system starting November'),('Spoke with','Denise, intake'),
    ('Call outcome','reached'),('Note',''),('Called (date)','2026-09-03'),('VA','JM')]
for k,v in ex:
    ca=gs.cell(r,1,k); cb=gs.cell(r,2,v)
    ca.font=Font(name='Arial',size=9,color='6B78A0'); cb.font=Font(name='Arial',size=9)
    cb.fill=PatternFill('solid',fgColor=YELLOW); cb.alignment=Alignment(wrap_text=True,vertical='top')
    r+=1
r+=1
r=line(r,'PROGRESS',' ',bold=True,fill=LAV)
L=get_column_letter(cols.index('verify_outcome')+1)
S=get_column_letter(cols.index('status')+1)
gs.cell(r,1,'Calls logged').font=Font(name='Arial',size=10,bold=True,color=NAVY)
gs.cell(r,2,f"=COUNTA('Call Sheet'!{L}{HR+1}:{L}{last})&\" of {len(rows)}\"").font=Font(name='Arial',size=10)
r+=1
gs.cell(r,1,'Reached').font=Font(name='Arial',size=10,bold=True,color=NAVY)
gs.cell(r,2,f"=COUNTIF('Call Sheet'!{L}{HR+1}:{L}{last},\"reached\")").font=Font(name='Arial',size=10)
r+=1
gs.cell(r,1,'Currently accepting').font=Font(name='Arial',size=10,bold=True,color=NAVY)
gs.cell(r,2,f"=COUNTIF('Call Sheet'!{S}{HR+1}:{S}{last},\"accepting\")").font=Font(name='Arial',size=10)
r+=1
gs.cell(r,1,'Reach rate').font=Font(name='Arial',size=10,bold=True,color=NAVY)
c=gs.cell(r,2,f"=IFERROR(COUNTIF('Call Sheet'!{L}{HR+1}:{L}{last},\"reached\")/COUNTA('Call Sheet'!{L}{HR+1}:{L}{last}),\"—\")")
c.font=Font(name='Arial',size=10,bold=True); c.number_format='0%'
r+=1
gs.cell(r,2,'This is the number the staffing model is guessing at (it assumes 60%). Measure it here.').font=Font(name='Arial',size=9,italic=True,color='6B78A0')

wb.active=0
out=sys.argv[2] if len(sys.argv)>2 else 'data/call-sheets/CornerHelp-Week1-CallSheet.xlsx'; wb.save(out); print('wrote', out)
